import json
from django.http import HttpResponse, JsonResponse
from system.models import Account
from django.views.decorators.csrf import csrf_exempt
from django.forms.models import model_to_dict
from django.db.models import Sum, Q
from base.models import Company, Code, IntroducingCompany, Excel
from system.models import Account
from django.db import transaction
from common.revisions import set_revisions, get_revisions
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import os
from openpyxl.styles import (
    Border, Side,
    Alignment,
	Font
)
@csrf_exempt
def CompanyCreateAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    try:
        Company.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['name'],
        )
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    except Company.DoesNotExist:
        Company.objects.create(**json_data)

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def CompanyModAPI(request,pk):
    json_data = json.loads(request.body.decode('utf-8'))
    #중복체크
    try:
        company = Company.objects.get(pk=pk)
        Company.objects.get(
            ~Q(pk = company.pk),
            introducing_company = company.introducing_company,
            name =json_data['name'],
        )
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    except Company.DoesNotExist:
        set_revisions(request, 'base', 'company', pk)
        # Company.objects.filter(pk=pk).update(**json_data)

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def CompanyDelAPI(request,pk):

    Company.objects.filter(pk=pk).delete()

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def GetCompaniesAPI(request):
    search_type = request.GET.get('search_type') or ''
    keyword = request.GET.get('keyword') or ''
    type = request.GET.get('type') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)


    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Company.objects.filter(introducing_company=introducing_company).order_by('-pk')

    if start_date != '':
        data_list = data_list.filter(
            created_dt__date__gte = start_date
        )

    if end_date != '':
        data_list = data_list.filter(
            created_dt__date__lte = end_date
        )

    if type != '':
        data_list = data_list.filter(
            type__icontains = type
        )
    if search_type == '':
        data_list = data_list.filter(
            Q(name__icontains = keyword)|
            Q(address__icontains = keyword)|
            Q(onwer_phone__icontains = keyword)|
            Q(line_of_business__icontains = keyword)|
            Q(type_of_business__icontains = keyword)|
            Q(email__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'type':
            q.add(Q(type__icontains = keyword), q.AND)
        elif search_type == 'address':
            q.add(Q(address__icontains = keyword), q.AND)
        elif search_type == 'onwer':
            q.add(Q(onwer__icontains = keyword), q.AND)
        elif search_type == 'onwer_phone':
            q.add(Q(onwer_phone__icontains = keyword), q.AND)
        elif search_type == 'line_of_business':
            q.add(Q(line_of_business__icontains = keyword), q.AND)
        elif search_type == 'type_of_business':
            q.add(Q(type_of_business__icontains = keyword), q.AND)
        elif search_type == 'email':
            q.add(Q(email__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
            
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','type','name','address','onwer_phone','onwer','line_of_business','type_of_business','email','memo'
    )[(page-1) * list_range :  page * list_range]

    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)

@csrf_exempt
def GetCompanyAPI(request,pk):
    try:
        return JsonResponse(model_to_dict(Company.objects.get(pk=pk)), safe=False)
    except Company.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    

@transaction.atomic
def CompanyListDelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    for data in data_list:
        company = Company.objects.get(
            id = data['id']
        )
        company.delete()
        
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@transaction.atomic
@csrf_exempt
def CompanyUploadAPI(request):
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    excel_file = request.FILES.getlist('excel_file')
    if len(excel_file) ==0:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    excel = Excel()
    excel.introducing_company = introducing_company
    excel.file_name = excel_file[0].name
    excel.upload_date = datetime.now()
    excel.kind = '업체'
    df = pd.read_excel(excel_file[0],sheet_name=0,header=0, dtype=str)
    df = df.replace(np.nan, '', regex=True)
    df = df.fillna("")
    items = df.to_dict('records')
    for item in items :
        if item['업체명'] == '':
            continue
        type = ''
        # TODO 업체구분 어떻게 들어가는지 체크
        if '매출' in item['업체구분'] and '매입' in item['업체구분']:
            type = '매입/매출처'
        elif '매출' in item['업체구분']:
            type = '매출처'
        elif '매입' in item['업체구분']:
            type = '매입처'

        company, flag = Company.objects.get_or_create(
            introducing_company = introducing_company,
            name = item['업체명']
        )
        company.type = type
        company.registration_num = item['사업자등록번호'] or ''
        company.address = item['업체주소'] or ''
        company.onwer = item['대표자'] or ''
        company.onwer_phone = item['대표번호'] or ''
        company.email = item['이메일'] or ''
        company.fax = item['팩스번호'] or ''
        company.type_of_business = item['종목'] or ''
        company.line_of_business = item['업태'] or ''
        company.manager_name = item['담당자이름'] or ''
        company.manager_phone = item['담당자연락처'] or ''
        company.memo = item['메모'] or ''
        company.save()
    excel.save()
    excel.excel_file = excel_file[0]
    excel.save()
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def CompanyExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    keyword = request.GET.get('keyword') or ''
    type = request.GET.get('type') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''

    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Company.objects.filter(introducing_company=introducing_company).order_by('name')

    if start_date != '':
        data_list = data_list.filter(
            created_dt__date__gte = start_date
        )

    if end_date != '':
        data_list = data_list.filter(
            created_dt__date__lte = end_date
        )

    if type != '':
        data_list = data_list.filter(
            type__icontains = type
        )
    if search_type == '':
        data_list = data_list.filter(
            Q(name__icontains = keyword)|
            Q(address__icontains = keyword)|
            Q(onwer_phone__icontains = keyword)|
            Q(line_of_business__icontains = keyword)|
            Q(type_of_business__icontains = keyword)|
            Q(email__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'type':
            q.add(Q(type__icontains = keyword), q.AND)
        elif search_type == 'address':
            q.add(Q(address__icontains = keyword), q.AND)
        elif search_type == 'onwer':
            q.add(Q(onwer__icontains = keyword), q.AND)
        elif search_type == 'onwer_phone':
            q.add(Q(onwer_phone__icontains = keyword), q.AND)
        elif search_type == 'line_of_business':
            q.add(Q(line_of_business__icontains = keyword), q.AND)
        elif search_type == 'type_of_business':
            q.add(Q(type_of_business__icontains = keyword), q.AND)
        elif search_type == 'email':
            q.add(Q(email__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
            
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','type','name','address','onwer_phone','onwer','line_of_business','type_of_business','email','memo'
    )

    path = './media/company'
    docx_title = 'export_company'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_company.xlsx')
    ws = wb["Sheet 1"]
    i = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)
    for data in data_list : 
        num = str(i)
        ws['A'+num].value = data['type']
        ws['B'+num].value = data['name'] or ''
        ws['C'+num].value = data['address'] or ''
        ws['D'+num].value = data['onwer_phone'] or ''
        ws['E'+num].value = data['onwer'] or ''
        ws['F'+num].value = data['line_of_business'] or ''
        ws['G'+num].value = data['email'] or ''
        ws['H'+num].value = data['memo'] or ''
     

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font
        ws['F'+num].font = font
        ws['G'+num].font = font
        ws['H'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
        ws['F'+num].border = thin_border
        ws['G'+num].border = thin_border
        ws['H'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
        ws['F'+num].alignment = alignment
        ws['G'+num].alignment = alignment
        ws['H'+num].alignment = alignment
       
        ws.row_dimensions[i].height = 40
     
        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')