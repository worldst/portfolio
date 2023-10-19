import json
from django.http import HttpResponse, JsonResponse
from system.models import Account
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q, F, Func, Value, CharField
from base.models import IntroducingCompany, Excel, Company, Code
from system.models import Account
from material.models import (
    Material, MaterialPurchase, MaterialPurchaseList, EnterRecord,
    MaterialUsedRecord
)
from django.forms.models import model_to_dict
from django.db import transaction
from common.revisions import set_revisions, get_revisions
from datetime import datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (
    Border, Side,
    Alignment,
	Font
)
import xlsxwriter
import qrcode
import os
import zipfile
import arrow
@csrf_exempt
def MaterialCreateAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    try:
        Material.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['name'],
        )
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    except Material.DoesNotExist:
        Material.objects.create(**json_data)

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def MaterialModAPI(request,pk):
    json_data = json.loads(request.body.decode('utf-8'))

    #중복체크
    data = Material.objects.get(pk=pk)
    try:
        Material.objects.get(
            ~Q(pk = data.pk),
            introducing_company = data.introducing_company,
            name = json_data['name'],
        )
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    except Material.DoesNotExist:
        set_revisions(request, 'material', 'material', pk, json_data)
        # Material.objects.filter(pk=pk).update(**json_data)


    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def MaterialDelAPI(request,pk):

    Material.objects.filter(pk=pk).delete()

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def GetMaterialsAPI(request):
    search_type = request.GET.get('search_type') or ''
    keyword = request.GET.get('keyword') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    
    data_list = Material.objects.filter(introducing_company=introducing_company).order_by('-pk')

    if start_date != '':
        data_list = data_list.filter(
            created_dt__date__gte = start_date
        )

    if end_date != '':
        data_list = data_list.filter(
            created_dt__date__lte = end_date
        )

    if search_type == '':
        data_list = data_list.filter(
            Q(name__icontains = keyword)|
            Q(unit_price__icontains = keyword)|
            Q(unit__icontains = keyword)|
            Q(stock__icontains = keyword)|
            Q(safety_stock__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'unit_price':
            q.add(Q(unit_price__icontains = keyword), q.AND)
        elif search_type == 'unit':
            q.add(Q(unit__icontains = keyword), q.AND)
        elif search_type == 'stock':
            q.add(Q(stock__icontains = keyword), q.AND)
        elif search_type == 'safety_stock':
            q.add(Q(safety_stock__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','name','unit_price','unit','stock','safety_stock'
    )[(page-1) * list_range :  page * list_range]
    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)


@csrf_exempt
def GetMaterialAPI(request,pk):
    try:
        return JsonResponse(model_to_dict(Material.objects.get(pk=pk)), safe=False)
    except Material.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')

@transaction.atomic
def MaterialListDelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    for data in data_list:
        material = Material.objects.get(
            id = data['id']
        )
        material.delete()
        
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@transaction.atomic
@csrf_exempt
def MaterialUploadAPI(request):
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    excel_file = request.FILES.getlist('excel_file')
    if len(excel_file) ==0:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')

    excel = Excel()
    excel.introducing_company = introducing_company
    excel.file_name = excel_file[0].name
    excel.upload_date = datetime.now()
    excel.kind = '자재'
    
    df = pd.read_excel(excel_file[0],sheet_name=0,header=0)
    df = df.replace(np.nan, '', regex=True)
    df = df.fillna("")
    items = df.to_dict('records')
    for item in items :
        if item['자재명'] == '':
            continue
        try:
            Material.objects.get(
                introducing_company = introducing_company,
                name = item['자재명'],
            )
        except Material.DoesNotExist:
            Material.objects.create(
                introducing_company = introducing_company,
                name = item['자재명'],
                unit_price = item['단가'] or 0 ,
                unit = item['단위'],
                stock = item['현재고'] or 0,
                safety_stock = item['안전재고'] or 0 ,
            )
    excel.save()
    excel.excel_file = excel_file[0]
    excel.save()
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def MaterialExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    keyword = request.GET.get('keyword') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    
    data_list = Material.objects.filter(introducing_company=introducing_company).order_by('-created_dt')

    if start_date != '':
        data_list = data_list.filter(
            created_dt__date__gte = start_date
        )

    if end_date != '':
        data_list = data_list.filter(
            created_dt__date__lte = end_date
        )

    if search_type == '':
        data_list = data_list.filter(
            Q(name__icontains = keyword)|
            Q(unit_price__icontains = keyword)|
            Q(unit__icontains = keyword)|
            Q(stock__icontains = keyword)|
            Q(safety_stock__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'unit_price':
            q.add(Q(unit_price__icontains = keyword), q.AND)
        elif search_type == 'unit':
            q.add(Q(unit__icontains = keyword), q.AND)
        elif search_type == 'stock':
            q.add(Q(stock__icontains = keyword), q.AND)
        elif search_type == 'safety_stock':
            q.add(Q(safety_stock__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','name','unit_price','unit','stock','safety_stock'
    )

    path = './media/material'
    docx_title = 'export_material'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_material.xlsx')
    ws = wb["Sheet 1"]
    i = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)
    for data in data_list : 
        num = str(i)
        ws['A'+num].value = data['name']
        ws['B'+num].value = data['unit_price'] or ''
        ws['C'+num].value = data['unit'] or ''
        ws['D'+num].value = data['stock'] or ''
        ws['E'+num].value = data['safety_stock'] or ''
     

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
       
        ws.row_dimensions[i].height = 40
     
        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')


@csrf_exempt
@transaction.atomic
def MaterialPurchaseCreateAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    material_purchase_list = json_data.pop('material_purchase_list')
    
    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')

    try:
        company = Company.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )
    except Company.DoesNotExist:
        company = Company.objects.create(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )

    json_data['company_id'] = company.pk
    # 주문 특이사항 변환
    special_note = ''
    for row in json_data['special_note']:
        if special_note == '':
            special_note = row
        else: 
            special_note += 'Ψ' + row
    json_data['special_note'] = special_note

    num = datetime.now().strftime('P%Y%m%d-')
    material_purchase = MaterialPurchase.objects.filter(
        introducing_company_id = json_data['introducing_company_id'],
        num__icontains = num 
    ).order_by('-num')[0:1]

    if len(material_purchase) != 0 :
        material_purchase = material_purchase.first()
        material_purchase_len = len(material_purchase.num.split('-'))
        num = num + str(int(material_purchase.num.split('-')[material_purchase_len-1]) + 1 ).zfill(3)
    else:
        num = num + '1'.zfill(3)

    json_data['num'] = num
    print(json_data)
    material_purchase = MaterialPurchase.objects.create(**json_data)

    for data in material_purchase_list:
        data['material_purchase_id'] = material_purchase.id
        MaterialPurchaseList.objects.create(**data)
      

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')
@csrf_exempt
def GetMaterialPurchasesAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    status = request.GET.get('status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = MaterialPurchase.objects.filter(introducing_company=introducing_company).order_by('-created_dt')


    if status != '':
        data_list = data_list.filter(
            status__in = status.split(",")
        )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'purchase_date':
            data_list = data_list.filter(
                purchase_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'purchase_date':
            data_list = data_list.filter(
                purchase_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','num','name','company_name','account__username','memo','status'
    ).annotate(
        created_dt  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        )
    )[(page-1) * list_range :  page * list_range]
    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)


@csrf_exempt
def MaterialPurchaseDelAPI(request,pk):

    MaterialPurchase.objects.filter(pk=pk).delete()

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def GetMaterialPurchaseAPI(request,pk):
    try:
        material_purchase = MaterialPurchase.objects.get(
            pk = pk
        )
      
        material_purchase_list = MaterialPurchaseList.objects.filter(
            material_purchase = material_purchase
        ).values('pk','material__pk','material__name','unit_price','cnt','price','upload_file')

        username = material_purchase.username
        material_purchase = (model_to_dict(material_purchase))
        material_purchase['username'] = username
        return JsonResponse(
            {
                'material_purchase': material_purchase,
                'material_purchase_list': list(material_purchase_list),
            }, safe=False)
    except MaterialPurchaseList.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')

@csrf_exempt
def MaterialPurchaseModAPI(request, pk):
    json_data = json.loads(request.body.decode('utf-8'))
    material_purchase_list = json_data.pop('material_purchase_list')

    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')

    try:
        company = Company.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )
    except Company.DoesNotExist:
        company = Company.objects.create(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )

    json_data['company_id'] = company.pk
    # 주문 특이사항 변환
    special_note = ''
    for row in json_data['special_note']:
        if special_note == '':
            special_note = row
        else: 
            special_note += 'Ψ' + row
    json_data['special_note'] = special_note
    set_revisions(request, 'material', 'materialpurchase', pk, json_data)
    material_purchase = MaterialPurchase.objects.get(
        pk = pk
    )

    if material_purchase.status == 'WAIT':
        MaterialPurchaseList.objects.filter(
            material_purchase = material_purchase
        ).delete()
        for data in material_purchase_list:
            data['material_purchase_id'] = material_purchase.id
            MaterialPurchaseList.objects.create(**data)

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@transaction.atomic
def MaterialPurchaseCancelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')

    for data in data_list:
        material_purchase = MaterialPurchase.objects.get(
            id = data['id']
        )
        if material_purchase.status != 'WAIT':
            # 상태가 대기가 아닌경우
            return HttpResponse(
                json.dumps({
                    'result': 'fail'
            }), content_type='application/json')
    for data in data_list:
        material_purchase = MaterialPurchase.objects.get(
            id = data['id']
        )
        material_purchase.cancel_memo = json_data['cancel_memo']
        material_purchase.status = 'CANCEL'
        material_purchase.save()
        
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@transaction.atomic
def MaterialPurchaseListDelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')

    for data in data_list:
        material_purchase = MaterialPurchase.objects.get(
            id = data['id']
        )
        if material_purchase.status != 'WAIT':
            # 상태가 대기가 아닌경우
            return HttpResponse(
                json.dumps({
                    'result': 'fail'
            }), content_type='application/json')
    for data in data_list:
        material_purchase = MaterialPurchase.objects.get(
            id = data['id']
        )
        material_purchase.delete()
        
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')


def GetMaterialPurchasesStatusChangeAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    for data in data_list:
        material_purchase = MaterialPurchase.objects.get(
            id = data['id']
        )
        if material_purchase.status == 'WAIT':
            material_purchase.start_dt = datetime.now()
        material_purchase.end_dt = datetime.now()
        material_purchase.status = 'COMPLT'
        material_purchase.save()
        material_purchase_list = MaterialPurchaseList.objects.filter(
            material_purchase = material_purchase
        )
        for data in material_purchase_list:
            if data.status == 'WAIT':
                data.start_dt = datetime.now()
            data.end_dt = datetime.now()
            data.status = 'COMPLT'
            data.save()
            if data.remain_cnt > 0 :
                EnterRecord.objects.create(
                    material_purchase_list = data,
                    cnt = data.remain_cnt,
                )
                material = data.material
                material.stock += data.remain_cnt
                data.complt_cnt += data.remain_cnt
                data.save()
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

def GetMaterialPurchaseStatusChangeAPI(request,pk):
    material_purchase = MaterialPurchase.objects.get(
        id = pk
    )
    if material_purchase.status == 'WAIT':
        material_purchase.start_dt = datetime.now()
    material_purchase.end_dt = datetime.now()
    material_purchase.status = 'COMPLT'
    material_purchase.save()
    material_purchase_list = MaterialPurchaseList.objects.filter(
        material_purchase = material_purchase
    )
    for data in material_purchase_list:
        if data.status == 'WAIT':
            data.start_dt = datetime.now()
        data.end_dt = datetime.now()
        data.status = 'COMPLT'
        data.save()
        if data.remain_cnt > 0 :
            EnterRecord.objects.create(
                material_purchase_list = data,
                cnt = data.remain_cnt,
            )
            material = data.material
            material.stock += data.remain_cnt
            material.save()
            data.complt_cnt += data.remain_cnt
            data.save()
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

# * 발주리스트 부분완료
def GetMaterialPurchaseListStatusChangeAPI(request,pk):
    json_data = json.loads(request.body.decode('utf-8'))
    cnt = json_data.pop('cnt')

    material_purchase_list = MaterialPurchaseList.objects.get(
        id = pk
    )
    if material_purchase_list.remain_cnt < cnt :
        cnt = material_purchase_list.remain_cnt
    if material_purchase_list.status == 'WAIT':
        material_purchase_list.start_dt = datetime.now()
    if cnt > 0 :
        material = material_purchase_list.material
        material.stock += cnt
        material.save()
        material_purchase_list.complt_cnt += cnt
        material_purchase_list.save()
        EnterRecord.objects.create(
            material_purchase_list = material_purchase_list,
            cnt = cnt,
        )
    # * 남은수량이 없을경우 완료 처리
    if material_purchase_list.remain_cnt <= 0 :
        material_purchase_list.status = 'COMPLT'
        material_purchase_list.end_dt = datetime.now()
    else:
        material_purchase_list.status = 'START'
    material_purchase_list.save()

    material_purchase = material_purchase_list.material_purchase
    if material_purchase.status == 'WAIT':
        material_purchase.start_dt = datetime.now()
        material_purchase.status = 'START'
    
    # * 발주리스트 전체완료 체크
    check = MaterialPurchaseList.objects.filter(
        ~Q(status = 'COMPLT'),
        material_purchase = material_purchase
    )
    if len(check) == 0 :
        material_purchase.end_dt = datetime.now()
        material_purchase.status = 'COMPLT'
    material_purchase.save()

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

def MaterialPurchaseExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    status = request.GET.get('status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    
    data_list = MaterialPurchase.objects.filter(introducing_company=introducing_company).order_by('-created_dt')


    if status != '':
        data_list = data_list.filter(
            status__in = status.split(",")
        )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'purchase_date':
            data_list = data_list.filter(
                purchase_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'purchase_date':
            data_list = data_list.filter(
                purchase_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','num','name','company_name','account__username','memo','status'
    ).annotate(
        created_dt  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        )
    )
    path = './media/material'
    docx_title = 'export_material_purchase'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_material_purchase.xlsx')
    ws = wb["Sheet 1"]
    i = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)
    for data in data_list : 
        num = str(i)
        status = '입고대기'
        if data['status'] == 'START':
            status = '부분입고'
        if data['status'] == 'CANCEL':
            status = '발주취소'
        if data['status'] == 'COMPLT':
            status = '입고완료'
        ws['A'+num].value = data['num'] or ''
        ws['B'+num].value = data['created_dt'] or ''
        ws['C'+num].value = data['name'] or ''
        ws['D'+num].value = data['company_name'] or ''
        ws['E'+num].value = data['account__username'] or ''
        ws['F'+num].value = data['memo'] or ''
        ws['G'+num].value = status
     

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font
        ws['F'+num].font = font
        ws['G'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
        ws['F'+num].border = thin_border
        ws['G'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
        ws['F'+num].alignment = alignment
        ws['G'+num].alignment = alignment
       
        ws.row_dimensions[i].height = 40
     
        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')

# * 자재입고기록
def MaterialEnterRecordAPI(request, pk):
    data_list = EnterRecord.objects.filter(
        material_purchase_list__material__pk = pk
    ).values(
        'cnt'
    ).order_by('-created_dt').annotate(
        enter_date  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
        material_name = F('material_purchase_list__material__name'),
        account_name = F('material_purchase_list__material_purchase__account__name'),
    )
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'data_list': list(data_list)
        }), content_type='application/json')

# * 자재사용기록
def MaterialUsedRecordAPI(request, pk):
    data_list = MaterialUsedRecord.objects.filter(
        material__pk = pk
    ).values(
        'cnt'
    ).order_by('-created_dt').annotate(
        used_date  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
        num = F('order__num'),
        material_name = F('material__name')
    )
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'data_list': list(data_list)
        }), content_type='application/json')

# * 자재 히스토리
def MaterialHistoryAPI(request, pk):
    return JsonResponse(
        {
            "data": get_revisions(request, 'material', 'material', pk)
        }
    )

@csrf_exempt
def MaterialPurchaseSearchAPI(request):
    num = request.GET.get('num')
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')

    try:
        material_purchase = MaterialPurchase.objects.get(
            introducing_company = introducing_company,
            num = num
        )
        
        if material_purchase.status == 'CANCEL':
            return HttpResponse(
                json.dumps({
                    'result': 'cancel',
            }), content_type='application/json')
        else:
            return HttpResponse(
                json.dumps({
                    'result': 'success',
                    'id' : material_purchase.pk
            }), content_type='application/json')
    except MaterialPurchase.DoesNotExist:
        # * 해당하는 주문이 존재 하지 않습니다.
        return HttpResponse(
            json.dumps({
                'result': 'no_order'
        }), content_type='application/json')

def WorkExcelCreate(pk):
    material_purchase = MaterialPurchase.objects.get(
        pk = pk
    )

    path = './media/material/work/'
    os.makedirs(path, exist_ok=True)

    wb = xlsxwriter.Workbook(path + material_purchase.num + '.xlsx')
    ws = wb.add_worksheet('자재발주서')

    ws.set_paper(9) # 용지크기설정
    ws.set_zoom(95)
    ws.fit_to_pages(1,1)
    ws.center_horizontally()# 가로 가운데
    # ws.center_vertically() # 세로 가운데
    ws.set_margins(0,0,0,0)

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=4,
        border=1,
    )
    qr.add_data(material_purchase.num)
    qr.make(fit=True)
    img = qr.make_image()
    img.save(path +'qrcode.png')
    ws.insert_image('B5', path + 'qrcode.png')

    def set_page(page):
        num = ( page -1 ) * 44 
        ws.print_area("B" + str(num + 5) + ":Y" + str(num + 44))
        ws.set_row(num, 4)
        ws.set_row(num + 1, 4)
        ws.set_row(num + 2, 4)
        ws.set_row(num + 3, 4)
        for i in range(num + 4, num + 44):
            ws.set_row(i, 20)
        
        ws.merge_range("B" + str(num + 5) + ":S" + str(num + 8), "자 재 발 주 서", fs_title)
        ws.merge_range("T" + str(num + 5) + ":U" + str(num + 5), '담당자', fs_top2)
        ws.merge_range("V" + str(num + 5) + ":W" + str(num + 5), '검토', fs_top2)
        ws.merge_range("X" + str(num + 5) + ":Y" + str(num + 5), '승인', fs_top2)
        ws.merge_range("T" + str(num + 6) + ":U" + str(num + 8), "", fs_top2)
        ws.merge_range("V" + str(num + 6) + ":W" + str(num + 8), "", fs_top2)
        ws.merge_range("X" + str(num + 6) + ":Y" + str(num + 8), "", fs_top2)

        ws.merge_range("B" + str(num + 9) + ":Y" + str(num + 9), "", fs_top3)

        ws.merge_range("B" + str(num + 10) + ":D" + str(num + 11), "발주번호", fs_mid)
        ws.merge_range("B" + str(num + 12) + ":D" + str(num + 13), "접수 담당자", fs_mid)
        ws.merge_range("B" + str(num + 14) + ":D" + str(num + 15), "발주건명", fs_mid)
        ws.merge_range("B" + str(num + 16) + ":Y" + str(num + 16), "공정 정보", fs_mid)
        ws.merge_range("B" + str(num + 17) + ":D" + str(num + 18), "구분", fs_mid)

        ws.merge_range("I" + str(num + 10) + ":J" + str(num + 11), "업체명", fs_mid)
        ws.merge_range("I" + str(num + 12) + ":J" + str(num + 13), "주문 내용", fs_mid)
        ws.merge_range("E" + str(num + 17) + ":S" + str(num + 18), "자재명", fs_mid)

        ws.merge_range("N" + str(num + 10) + ":P" + str(num + 11), "발주 생성일", fs_mid)
        ws.merge_range("T" + str(num + 10) + ":V" + str(num + 11), "예상 요청일", fs_mid)
        ws.merge_range("T" + str(num + 12) + ":V" + str(num + 13), "", fs_mid)
        ws.merge_range("T" + str(num + 17) + ":V" + str(num + 18), "단위", fs_mid)
        ws.merge_range("W" + str(num + 17) + ":Y" + str(num + 18), "수량", fs_mid)

        ws.merge_range("E" + str(num + 10) + ":H" + str(num + 11), material_purchase.num, fs_mid2)
        ws.merge_range("E" + str(num + 12) + ":H" + str(num + 13), material_purchase.username or '', fs_mid2)
        ws.merge_range("E" + str(num + 14) + ":Y" + str(num + 15), material_purchase.company_name or '', fs_mid2)

        ws.merge_range("K" + str(num + 10) + ":M" + str(num + 11), material_purchase.name or '', fs_mid2)
        ws.merge_range("K" + str(num + 12) + ":S" + str(num + 13), material_purchase.special_note.replace('Ψ',','), fs_mid2)

        ws.merge_range("Q" + str(num + 10) + ":S" + str(num + 11), str(material_purchase.purchase_date), fs_mid2)
        ws.merge_range("W" + str(num + 10) + ":Y" + str(num + 11), str(material_purchase.due_date), fs_mid2)
        ws.merge_range("W" + str(num + 12) + ":Y" + str(num + 13), '', fs_mid2)
        
        ws.merge_range("B"+str(num + 44)+":D" + str(num + 44), '메모', fs_mid)
        ws.merge_range("E"+str(num + 44)+":Y" + str(num + 44), material_purchase.memo, fs_bot)

        for i in  range(num + 19, num + 44) :
            ws.merge_range("B"+str(i)+":D" + str(i), '', fs_bot)
            ws.merge_range("E"+str(i)+":S" + str(i), '', fs_bot)
            ws.merge_range("T"+str(i)+":V" + str(i), '', fs_bot)
            ws.merge_range("W"+str(i)+":Y" + str(i), '', fs_bot)
        
        ws.insert_image('B' + str(num + 5), path + 'qrcode.png',{'x_scale': 1, 'y_scale': 1.1})

    ws.set_column('A:A', 0.36)
    ws.set_column('B:B', 2.7)
    ws.set_column('C:C', 2.7)
    ws.set_column('D:D', 2.7)
    ws.set_column('E:E', 4.2)
    ws.set_column('F:F', 4.2)
    ws.set_column('G:G', 4.2)
    ws.set_column('H:H', 3.5)
    ws.set_column('I:I', 3.5)
    ws.set_column('J:J', 3.5)
    ws.set_column('K:K', 3.5)
    ws.set_column('L:L', 3.5)
    ws.set_column('M:M', 3.5)
    ws.set_column('N:N', 2.7)
    ws.set_column('O:O', 2.7)
    ws.set_column('P:P', 2.7)
    ws.set_column('Q:Q', 3.2)
    for i in range(82, 90):
        ch = chr(i)
        ws.set_column(ch+":"+ch, 3.8)
    ws.set_column('Z:Z', 0.4)

    fs_title = wb.add_format({
        'bold' : 1,
        'font_name' : '굴림',
        'font_size': 30,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_top = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'font_size' : 10,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_top2 = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_top3 = wb.add_format({
        'font_name' : '굴림',
        'font_size' : 11,
        'shrink' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_mid = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'font_size' : 12,
        'shrink' : True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#e5e5e5',
    })

    fs_mid2 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'border' : 1,
        'font_size' : 12,
        'shrink' : True,
        'align': 'left',
        'valign': 'vcenter',
    })

    fs_mid3 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'top' : 1,
        'bottom' : 1,
        'shrink' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_bot = wb.add_format({
        'font_name' : '굴림',
        'font_size' : 8,
        'border' : 1,
        'font_size' : 13,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_bot2 = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_bot3 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'font_size' : 11,
        'border' : 1,
        'text_wrap' : True,
        'align': 'left',
        'valign': 'top',
    })

    fs_bot4 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'font_size' : 11,
        'border' : 1,
        'text_wrap' : True,
        'align': 'left',
        'valign': 'vcenter',
    })

    data_len = len(MaterialPurchaseList.objects.filter(material_purchase=material_purchase))
    pages = int( data_len / 25 )
    if data_len - pages * 25 != 0 :
        pages = pages + 1
    if data_len == 0 :
        pages = 1

    for i in range(1, pages + 1):
        set_page(i)
    num = 19
    count = 0
    for data in  MaterialPurchaseList.objects.filter(material_purchase=material_purchase) :
        sort = '자재'
        ws.write("B"+str(num), sort, fs_bot)
        ws.write("E"+str(num), data.material.name, fs_bot)
        ws.write("T"+str(num), data.unit, fs_bot)
        ws.write("W"+str(num), data.cnt, fs_bot)
        count += 1
        if count == 25 : 
            num += 20
            count = 0 
        else:
            num += 1
    
    wb.close()

    return material_purchase.num

def MaterialPurchasesWorkExcelDownloadAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    file_list = []
    path = './media/material/work/'
    os.makedirs(path, exist_ok=True)
    for data in data_list: 
        file_name = WorkExcelCreate(data['id'])
        file_list.append(path + '/' + file_name + '.xlsx')
    if len(file_list) == 1 :
        return HttpResponse(
            json.dumps({
                'result': 'success',
                'sort' : 'xlsx',
                'file_name' : file_name
        }), content_type='application/json')
    else:
        date_te=arrow.now().format("YY-MM-DD-HHmm")
        zip_subdir = date_te
        zip_filename = "%s.zip" % zip_subdir
        os.makedirs(path, exist_ok=True)
        zf = zipfile.ZipFile(path+'/'+date_te+".zip", "w")

        for fpath in file_list:
            fdir, fname = os.path.split(fpath)
            zip_path = os.path.join(zip_subdir, fname)
            zf.write(fpath,zip_path)
        zf.close()
        file_name = str(zip_filename.split('.zip')[0])
        return HttpResponse(
            json.dumps({
                'result': 'success',
                'sort' : 'zip',
                'file_name' : file_name
        }), content_type='application/json')
    
def MaterialPurchaseWorkExcelDownloadAPI(request):
    pk = request.POST.get('pk')

    file_name = WorkExcelCreate(pk)
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'file_name' : file_name
    }), content_type='application/json')