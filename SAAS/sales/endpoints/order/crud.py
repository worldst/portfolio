import json
from django.http import HttpResponse, JsonResponse
from system.models import Account
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q, F, Func, Value, CharField, Subquery, OuterRef
from base.models import Code, Company, IntroducingCompany
from system.models import Account
from django.forms.models import model_to_dict
from sales.models import (
    Order, OrderProcess, OrderMaterialOutsourcing, OrderFile, 
    ShipmentData, ShipmentFile
)
from datetime import datetime
from common.get_literal_eval import get_literal_eval
from ast import literal_eval

from django.db import transaction
from common.revisions import set_revisions, get_revisions
from material.models import MaterialUsedRecord
import openpyxl
import os
from openpyxl.styles import (
    Border, Side,
    Alignment,
	Font
)
import xlsxwriter
import qrcode
import os
import zipfile
import arrow
@csrf_exempt
@transaction.atomic
def OrderCreateAPI(request):
    json_data = json.loads(request.POST.get('obj'))
    order_process = json_data.pop('order_process')
    order_material_outsourcing = json_data.pop('order_material_outsourcing')
    
    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')

    try:
        company = Company.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )
    except Company.DoesNotExist:
        company = Company.objects.create(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )

    json_data['company_id'] = company.pk
    # 주문 특이사항 변환
    special_note = ''
    for row in json_data['special_note']:
        if special_note == '':
            special_note = row
        else: 
            special_note += 'Ψ' + row
    json_data['special_note'] = special_note

    num = datetime.now().strftime('%Y%m%d-')
    order = Order.objects.filter(
        introducing_company_id = json_data['introducing_company_id'],
        num__icontains = num 
    ).order_by('-num')[0:1]

    if len(order) != 0 :
        order = order.first()
        order_len = len(order.num.split('-'))
        num = num + str(int(order.num.split('-')[order_len-1]) + 1 ).zfill(3)
    else:
        num = num + '1'.zfill(3)

    json_data['num'] = num
    order = Order.objects.create(**json_data)

    for data in order_process:
        OrderProcess.objects.create(
            order = order,
            process = data['process'],
            ordering = data['ordering'],
            is_active = data['is_active'],
            price = data['price']
        )

    file_idx = 0
    for data in order_material_outsourcing:
        obj = OrderMaterialOutsourcing.objects.create(
            order = order,
            material_id = data['material_id'],
            sort = data['sort'],
            name = data['name'],
            unit_price = data['unit_price'],
            unit = data['unit'],
            cnt = data['cnt'],
            price = data['price'],
        )
        if data['is_file'] is 'Y':
            obj.upload_file = request.FILES.getlist('upload_file')[file_idx]
            obj.save()
            file_idx += 1

    for upload_file in request.FILES.getlist('order_file'):
        OrderFile.objects.create(
            order = order,
            upload_file = upload_file
        )

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def OrderDelAPI(request,pk):

    Order.objects.filter(pk=pk).delete()

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def GetOrdersAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    status = request.GET.get('status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Order.objects.filter(introducing_company=introducing_company).order_by('-created_dt')


    if status != '':
        data_list = data_list.filter(
            status__in = status.split(",")
        )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(pjt_name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(pjt_name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','num','pjt_name','company_name','account__username','memo','status','shipment_status','is_export'
    ).annotate(
        created_dt  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        )
    )[(page-1) * list_range :  page * list_range]
    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)

@csrf_exempt
def GetOrderAPI(request,pk):
    try:
        order = Order.objects.get(
            pk = pk
        )
        order_process = OrderProcess.objects.filter(
            order = order
        ).values()

        order_file = OrderFile.objects.filter(
            order = order
        ).values('pk','upload_file')

        order_material_outsourcing = OrderMaterialOutsourcing.objects.filter(
            order = order
        ).values('pk','sort','name','unit_price','cnt','price','upload_file')

        username = order.username
        order = (model_to_dict(order))
        order['username'] = username
        return JsonResponse(
            {
                'order': order,
                'order_process': list(order_process),
                'order_file': list(order_file),
                'order_material_outsourcing': list(order_material_outsourcing),
            }, safe=False)
    except Order.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')

@csrf_exempt
def OrderModAPI(request, pk):
    json_data = json.loads(request.POST.get('obj'))
    order_process = json_data.pop('order_process')

    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')

    try:
        company = Company.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )
    except Company.DoesNotExist:
        company = Company.objects.create(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )

    json_data['company_id'] = company.pk
    # 주문 특이사항 변환
    special_note = ''
    for row in json_data['special_note']:
        if special_note == '':
            special_note = row
        else: 
            special_note += 'Ψ' + row
    json_data['special_note'] = special_note
    # Order.objects.filter(pk=pk).update(**json_data)
    set_revisions(request, 'sales', 'order', pk, json_data)
    order = Order.objects.get(
        pk = pk
    )
    #TODO 완료시점에서 변경할시 자재 원복(원래 False인경우)
    for data in order_process:
        order_process = OrderProcess.objects.get(
            pk = data['pk']
        )
        order_process.process = data['process']
        order_process.ordering = data['ordering']
        order_process.is_active = data['is_active']
        order_process.price = data['price']
        order_process.save()

    # 수주 완료 처리가 되지 않은 건에 대해서만 자재 변경 가능
    if order.status != 'COMPLT':
        OrderMaterialOutsourcing.objects.filter(
            order = order
        ).delete()

        order_material_outsourcing = json_data.pop('order_material_outsourcing')

        file_idx = 0
        for data in order_material_outsourcing:
            obj = OrderMaterialOutsourcing.objects.create(
                order = order,
                material_id = data['material_id'],
                sort = data['sort'],
                name = data['name'],
                unit_price = data['unit_price'],
                unit = data['unit'],
                cnt = data['cnt'],
                price = data['price'],
            )
            if data['is_file'] is 'Y':
                obj.upload_file = request.FILES.getlist('upload_file')[file_idx]
                obj.save()
                file_idx += 1

    if order.status == 'COMPLT':
        # * 주문이 완료일때, 공정이 is_active True 될 경우
        check = OrderProcess.objects.filter(
            ~Q(status = 'COMPLT'),
            is_active=True,
            order = order
        )
        if len(check) != 0 :
            order.status = 'START'
            order.end_dt = None
            order.save()
            material_used_record =  MaterialUsedRecord.objects.filter(
                order = order
            )
            for data in material_used_record:
                material = data.material
                material.stock += data.cnt
                material.save()
                data.delete()
    

    OrderFile.objects.filter(
        order = order
    ).delete()
    
    for upload_file in request.FILES.getlist('order_file'):
        OrderFile.objects.create(
            order = order,
            upload_file = upload_file
        )

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@transaction.atomic
def OrderCancelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')

    for data in data_list:
        order = Order.objects.get(
            id = data['id']
        )
        if order.status == 'COMPLT' or order.shipment_status != 'WAIT':
            # 주문이 완료 상태이거나, 출하가 된 상태일때
            return HttpResponse(
                json.dumps({
                    'result': 'fail'
            }), content_type='application/json')
    for data in data_list:
        order = Order.objects.get(
            id = data['id']
        )
        order.cancel_memo = json_data['cancel_memo']
        order.status = 'CANCEL'
        order.save()
        
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@transaction.atomic
def OrderListDelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')

    for data in data_list:
        order = Order.objects.get(
            id = data['id']
        )
        if order.status == 'COMPLT' or order.shipment_status != 'WAIT':
            # 주문이 완료 상태이거나, 출하가 된 상태일때
            return HttpResponse(
                json.dumps({
                    'result': 'fail'
            }), content_type='application/json')
    for data in data_list:
        order = Order.objects.get(
            id = data['id']
        )
        order.delete()
        
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')


@csrf_exempt
def GetOrderShipmentsAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    shipment_status = request.GET.get('shipment_status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Order.objects.filter(introducing_company=introducing_company).order_by('-pk')


    if shipment_status != '':
        if shipment_status == 'DELAY':
            data_list = data_list.filter(
                ~Q(shipment_status = 'COMPLT'),
                due_date__lt = datetime.now().date()
            )
        else:
            data_list = data_list.filter(
                shipment_status = shipment_status
            )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(pjt_name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(pjt_name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','num','pjt_name','company_name','account__username','memo','shipment_status','is_export'
    ).annotate(
        due_date = Func(
            F('due_date'),
            Value('YYYY-MM-DD'),
            function='to_char',
            output_field=CharField()
        ),
        img_count = Subquery(
            ShipmentFile.objects.filter(
                order=OuterRef('pk'), sort='img'
            ).annotate(
                count=Func(F('pk'), function='Count'
            )).values('count')
        )
    )[(page-1) * list_range :  page * list_range]
    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)


@transaction.atomic
def GetOrderShipmentsStatusChangeAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    shipment_data = json_data.pop('shipment_data')
    shipment_status = shipment_data.pop('shipment_status')
    for data in data_list:
        order = Order.objects.get(
            id = data['id']
        )
        if order.shipment_status == 'WAIT':
            order.shipment_start_dt = datetime.now()
        if shipment_status == 'COMPLT':
            order.shipment_end_dt = datetime.now()
        order.shipment_status = shipment_status
        order.save()
        shipment_data['order_id'] = order.id
        ShipmentData.objects.create(**shipment_data)
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@transaction.atomic
@csrf_exempt
def GetOrderShipmentStatusChangeAPI(request, pk):
    json_data = json.loads(request.body.decode('utf-8'))
    shipment_status = json_data['shipment_status']
    order = Order.objects.get(
        pk = pk
    )
    if order.shipment_status == 'WAIT':
        order.shipment_start_dt = datetime.now()
    if shipment_status == 'COMPLT':
        order.shipment_end_dt = datetime.now()
    order.shipment_status = shipment_status
    order.save()
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')


@csrf_exempt
def GetOrderShipmentAPI(request,pk):
    try:
        order = Order.objects.get(
            pk = pk
        )

        order_material_outsourcing = OrderMaterialOutsourcing.objects.filter(
            order = order
        ).values('pk','sort','name','unit_price','cnt','price','upload_file')

        shipment_file = ShipmentFile.objects.filter(
            order = order
        ).values('pk','upload_file')

        shipment_data = ShipmentData.objects.filter(
            order = order
        ).values('pk','account__username','memo','method','name','phone_number','car_number')

        username = order.username
        order = (model_to_dict(order))
        order['username'] = username
        return JsonResponse(
            {
                'order': order,
                'order_material_outsourcing': list(order_material_outsourcing),
                'shipment_file': list(shipment_file),
                'shipment_data': list(shipment_data),
            }, safe=False)
    except Order.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')

@csrf_exempt
def OrderShipmentModAPI(request, pk):
    order = Order.objects.get(
        pk = pk
    )
    json_data = json.loads(request.POST.get('obj'))
    shipment_data = json_data.pop('shipment_data')

    ShipmentData.objects.filter(
        order = order
    ).delete()

    for data in shipment_data:
        ShipmentData.objects.create(
            order = order,
            account_id = data['account_id'],
            memo = data['memo'],
            method = data['method'],
            name = data['name'],
            phone_number = data['phone_number'],
            car_number = data['car_number']
        )
    ShipmentFile.objects.filter(
        order = order
    ).delete()
    for upload_file in request.FILES.getlist('shipment_file'):
        ShipmentFile.objects.create(
            order = order,
            upload_file = upload_file,
            sort = 'file'
        )
    for upload_file in request.FILES.getlist('shipment_img'):
        ShipmentFile.objects.create(
            order = order,
            upload_file = upload_file,
            sort = 'img'
        )

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def OrderDashboardRefreshAPI(request):
    data_list = []
    #페이지당 데이터 기본 100개
    list_range = request.GET.get('list_range') or 2
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)

    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    order_list = Order.objects.filter(
        introducing_company = introducing_company,
        status__in = ['WAIT','START']
    ).order_by('due_date','pk')[(page-1) * list_range :  page * list_range]
    process_list = Code.objects.filter(
        introducing_company = introducing_company,
        sort = 'process',
        # is_active = True
    ).order_by('ordering')
    for order in order_list:
        process_data = []
        for process in process_list:
            try:
                data = OrderProcess.objects.get(
                    order = order,
                    process = process.name,
                    is_active = True
                )
                
                process_data.append(
                    dict(
                        pk = data.pk,
                        name = data.process,
                        status = data.status,
                        worker_id = data.worker.id if data.worker else '',
                        worker = data.worker.name if data.worker else '',
                        last_equipment = data.last_equipment.id if data.last_equipment else ''
                    )
                )
            except:
                process_data.append(
                    dict()
                )
                continue

        account = ''
        if order.account != None:
            account = order.account.name

        data_list.append(
            dict(
                pk = order.pk,
                due_date = str(order.due_date),
                pjt_name = order.pjt_name,
                complt_rate = order.complt_rate,
                account = account,
                process_data = process_data
            )
        )
    total = len(Order.objects.filter(
        introducing_company_id = request.session['introducing_company_id'],
        status__in = ['WAIT','START']
    ))
    wait = len(Order.objects.filter(
        introducing_company_id = request.session['introducing_company_id'],
        status__in = ['WAIT']
    ))
    start = len(Order.objects.filter(
        introducing_company_id = request.session['introducing_company_id'],
        status__in = ['START']
    ))
    delay = len(Order.objects.filter(
        introducing_company_id = request.session['introducing_company_id'],
        status__in = ['WAIT','START'],
        due_date__gt = datetime.now().date()
    ))
    urgent = len(Order.objects.filter(
        introducing_company_id = request.session['introducing_company_id'],
        special_note__icontains = '긴급',
        status__in = ['WAIT','START'],
    ))
    total_process = len(OrderProcess.objects.filter(
        order__status__in = ['WAIT','START'],
        is_active = True
    ))
    complt_process = len(OrderProcess.objects.filter(
        status__in = ['COMPLT'],
        order__status__in = ['WAIT','START'],
        is_active = True,
    ))
    if total_process != 0 and complt_process != 0 :
        complt_rate = int(complt_process / total_process * 100)
    else:
        complt_rate = 0

    data_len = len(Order.objects.filter(
        introducing_company = introducing_company,
        status__in = ['WAIT','START']
    ))
    print(data_len)
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'total' : total,
            'wait' : wait,
            'start' : start,
            'delay' : delay,
            'urgent' : urgent,
            'complt_rate' : complt_rate,
            'data_len' : data_len,
            'data_list' : data_list,
            'code_list' : list(process_list.values('id','name')),
    }), content_type='application/json')

@csrf_exempt
def OrderExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    status = request.GET.get('status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Order.objects.filter(introducing_company=introducing_company).order_by('-created_dt')


    if status != '':
        data_list = data_list.filter(
            status__in = status.split(",")
        )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(pjt_name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(pjt_name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','num','pjt_name','company_name','account__username','memo','status','shipment_status'
    ).annotate(
        created_dt  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        )
    )
    path = './media/sales/order'
    docx_title = 'export_order'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_order.xlsx')
    ws = wb["Sheet 1"]
    i = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)
    for data in data_list : 
        num = str(i)
        status = '생산대기'
        if data['status'] == 'START':
            status = '생산시작'
        if data['status'] == 'PAUSE':
            status = '일시정지'
        if data['status'] == 'CANCEL':
            status = '수주취소'
        if data['status'] == 'PART':
            status = '부분완료'
        if data['status'] == 'COMPLT':
            status = '수주완료'
        shipment_status = '출하대기'
        if data['shipment_status'] == 'START':
            shipment_status = '부분출하'
        if data['shipment_status'] == 'COMPLT':
            status = '출하완료'
        ws['A'+num].value = data['num']
        ws['B'+num].value = data['created_dt']
        ws['C'+num].value = data['pjt_name'] or ''
        ws['D'+num].value = data['company_name'] or ''
        ws['E'+num].value = data['account__username'] or ''
        ws['F'+num].value = data['memo'] or ''
        ws['G'+num].value = status
        ws['H'+num].value = shipment_status

     

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font
        ws['F'+num].font = font
        ws['G'+num].font = font
        ws['H'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
        ws['F'+num].border = thin_border
        ws['G'+num].border = thin_border
        ws['H'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
        ws['F'+num].alignment = alignment
        ws['G'+num].alignment = alignment
        ws['H'+num].alignment = alignment
       
        ws.row_dimensions[i].height = 40
     
        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')

@csrf_exempt
def OrderShipmentExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    shipment_status = request.GET.get('shipment_status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Order.objects.filter(introducing_company=introducing_company).order_by('-created_dt')


    if shipment_status != '':
        data_list = data_list.filter(
            shipment_status = shipment_status
        )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(pjt_name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(pjt_name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','num','pjt_name','company_name','account__username','memo','shipment_status'
    ).annotate(
        due_date  = Func(
            F('due_date'),
            Value('YYYY-MM-DD'),
            function='to_char',
            output_field=CharField()
        )
    )
    path = './media/sales/shipment'
    docx_title = 'export_shipment'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_shipment.xlsx')
    ws = wb["Sheet 1"]
    i = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)
    for data in data_list : 
        num = str(i)
        # TODO 출하상태 확인해야함
        shipment_status = '출하대기'
        if data['shipment_status'] == 'START':
            shipment_status = '부분출하'
        if data['shipment_status'] == 'COMPLT':
            shipment_status = '출하완료'
        ws['A'+num].value = data['num']
        ws['B'+num].value = data['due_date']
        ws['C'+num].value = data['pjt_name'] or ''
        ws['D'+num].value = data['company_name'] or ''
        ws['E'+num].value = data['account__username'] or ''
        ws['F'+num].value = data['memo'] or ''
        ws['G'+num].value = shipment_status

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font
        ws['F'+num].font = font
        ws['G'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
        ws['F'+num].border = thin_border
        ws['G'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
        ws['F'+num].alignment = alignment
        ws['G'+num].alignment = alignment
       
        ws.row_dimensions[i].height = 40
     
        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')


@csrf_exempt
def OrderShipmentPhotoCreateAPI(request, pk):
    order = Order.objects.get(
        pk = pk
    )
    for upload_file in request.FILES.getlist('shipment_file'):
        ShipmentFile.objects.create(
            order = order,
            upload_file = upload_file
        )
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

def WorkExcelCreate(pk):
    order = Order.objects.get(
        pk = pk
    )

    path = './media/sales/work/'
    os.makedirs(path, exist_ok=True)

    wb = xlsxwriter.Workbook(path + order.num + '.xlsx')
    ws = wb.add_worksheet('작업지시서')

    ws.set_paper(9) # 용지크기설정
    ws.set_zoom(95)
    ws.fit_to_pages(1,1)
    ws.center_horizontally()# 가로 가운데
    # ws.center_vertically() # 세로 가운데
    ws.set_margins(0,0,0,0)

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=4,
        border=1,
    )
    qr.add_data(order.num)
    qr.make(fit=True)
    img = qr.make_image()
    img.save(path +'qrcode.png')

    def set_page(page):
        num = ( page -1 ) * 46 
        ws.print_area("B" + str(num + 5) + ":Y" + str(num + 46))
        ws.set_row(num, 4)
        ws.set_row(num + 1, 4)
        ws.set_row(num + 2, 4)
        ws.set_row(num + 3, 4)
        for i in range(num + 4, num + 46):
            ws.set_row(i, 20)
        
        ws.merge_range("B" + str(num + 5) + ":S" + str(num + 8), "작 업 지 시 서", fs_title)
        ws.merge_range("T" + str(num + 5) + ":U" + str(num + 5), '담당자', fs_top2)
        ws.merge_range("V" + str(num + 5) + ":W" + str(num + 5), '검토', fs_top2)
        ws.merge_range("X" + str(num + 5) + ":Y" + str(num + 5), '승인', fs_top2)
        ws.merge_range("T" + str(num + 6) + ":U" + str(num + 8), "", fs_top2)
        ws.merge_range("V" + str(num + 6) + ":W" + str(num + 8), "", fs_top2)
        ws.merge_range("X" + str(num + 6) + ":Y" + str(num + 8), "", fs_top2)

        ws.merge_range("B" + str(num + 9) + ":Y" + str(num + 9), "", fs_top3)

        ws.merge_range("B" + str(num + 10) + ":D" + str(num + 11), "주문번호", fs_mid)
        ws.merge_range("B" + str(num + 12) + ":D" + str(num + 13), "접수 담당자", fs_mid)
        ws.merge_range("B" + str(num + 14) + ":D" + str(num + 15), "프로젝트명", fs_mid)
        ws.merge_range("B" + str(num + 16) + ":D" + str(num + 17), "공정 정보", fs_mid)
        ws.merge_range("B" + str(num + 18) + ":Y" + str(num + 18), "", fs_mid)
        ws.merge_range("B" + str(num + 19) + ":D" + str(num + 20), "구분", fs_mid)

        ws.merge_range("I" + str(num + 10) + ":J" + str(num + 11), "업체명", fs_mid)
        ws.merge_range("I" + str(num + 12) + ":J" + str(num + 13), "주문 특이사항", fs_mid)
        ws.merge_range("E" + str(num + 19) + ":S" + str(num + 20), "자재명 / 외주명", fs_mid)

        ws.merge_range("N" + str(num + 10) + ":P" + str(num + 11), "주문일자", fs_mid)
        ws.merge_range("T" + str(num + 10) + ":V" + str(num + 11), "납품 예정 일자", fs_mid)
        ws.merge_range("T" + str(num + 12) + ":V" + str(num + 13), "배송 방법", fs_mid)
        ws.merge_range("T" + str(num + 19) + ":V" + str(num + 20), "단위", fs_mid)
        ws.merge_range("W" + str(num + 19) + ":Y" + str(num + 20), "수량", fs_mid)

        ws.merge_range("E" + str(num + 10) + ":H" + str(num + 11), order.num, fs_mid2)
        ws.merge_range("E" + str(num + 12) + ":H" + str(num + 13), order.username or '', fs_mid2)
        ws.merge_range("E" + str(num + 14) + ":Y" + str(num + 15), order.pjt_name or '', fs_mid2)
        ws.merge_range("E" + str(num + 16) + ":Y" + str(num + 17), order.process_text or '', fs_mid2)

        ws.merge_range("K" + str(num + 10) + ":M" + str(num + 11), order.company_name or '', fs_mid2)
        ws.merge_range("K" + str(num + 12) + ":S" + str(num + 13), order.special_note.replace('Ψ',','), fs_mid2)

        ws.merge_range("Q" + str(num + 10) + ":S" + str(num + 11), str(order.registration_date), fs_mid2)
        ws.merge_range("W" + str(num + 10) + ":Y" + str(num + 11), str(order.due_date), fs_mid2)
        ws.merge_range("W" + str(num + 12) + ":Y" + str(num + 13), order.delivery_method, fs_mid2)
        
        ws.merge_range("B"+str(num + 46)+":D" + str(num + 46), '메모', fs_mid)
        ws.merge_range("E"+str(num + 46)+":Y" + str(num + 46), order.memo, fs_bot)

        for i in  range(num + 21, num + 46) :
            ws.merge_range("B"+str(i)+":D" + str(i), '', fs_bot)
            ws.merge_range("E"+str(i)+":S" + str(i), '', fs_bot)
            ws.merge_range("T"+str(i)+":V" + str(i), '', fs_bot)
            ws.merge_range("W"+str(i)+":Y" + str(i), '', fs_bot)
        
        ws.insert_image('B' + str(num + 5), path + 'qrcode.png',{'x_scale': 1, 'y_scale': 1.1})


    ws.set_column('A:A', 0.36)
    ws.set_column('B:B', 2.7)
    ws.set_column('C:C', 2.7)
    ws.set_column('D:D', 2.7)
    ws.set_column('E:E', 4.2)
    ws.set_column('F:F', 4.2)
    ws.set_column('G:G', 4.2)
    ws.set_column('H:H', 3.5)
    ws.set_column('I:I', 3.5)
    ws.set_column('J:J', 3.5)
    ws.set_column('K:K', 3.5)
    ws.set_column('L:L', 3.5)
    ws.set_column('M:M', 3.5)
    ws.set_column('N:N', 2.7)
    ws.set_column('O:O', 2.7)
    ws.set_column('P:P', 2.7)
    ws.set_column('Q:Q', 3.2)
    for i in range(82, 90):
        ch = chr(i)
        ws.set_column(ch+":"+ch, 3.8)
    ws.set_column('Z:Z', 0.4)

    fs_title = wb.add_format({
        'bold' : 1,
        'font_name' : '굴림',
        'font_size': 30,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_top = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'font_size' : 10,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_top2 = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_top3 = wb.add_format({
        'font_name' : '굴림',
        'font_size' : 11,
        'shrink' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_mid = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'font_size' : 12,
        'shrink' : True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#e5e5e5',
    })

    fs_mid2 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'border' : 1,
        'font_size' : 12,
        'shrink' : True,
        'align': 'left',
        'valign': 'vcenter',
    })

    fs_mid3 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'top' : 1,
        'bottom' : 1,
        'shrink' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_bot = wb.add_format({
        'font_name' : '굴림',
        'font_size' : 13,
        'border' : 1,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_bot2 = wb.add_format({
        'font_name' : '굴림',
        'border' : 1,
        'text_wrap' : True,
        'align': 'center',
        'valign': 'vcenter',
    })

    fs_bot3 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'font_size' : 11,
        'border' : 1,
        'text_wrap' : True,
        'align': 'left',
        'valign': 'top',
    })

    fs_bot4 = wb.add_format({
        'font_name' : '굴림',
        'bold' : 1,
        'font_size' : 11,
        'border' : 1,
        'text_wrap' : True,
        'align': 'left',
        'valign': 'vcenter',
    })

    data_len = len(OrderMaterialOutsourcing.objects.filter(order=order))

    pages = int( data_len / 25 )

    if data_len - pages * 25 != 0 :
        pages = pages + 1
    if data_len == 0 :
        pages = 1
    for i in range(1, pages + 1):
        set_page(i)

    num = 21
    count = 0
    for data in  OrderMaterialOutsourcing.objects.filter(order=order) :
        
        sort = '외주'
        if data.sort == 'material':
            sort = '자재'
        ws.write("B"+str(num), sort, fs_bot)
        ws.write("E"+str(num), data.name, fs_bot)
        ws.write("T"+str(num), data.unit, fs_bot)
        ws.write("W"+str(num), data.cnt, fs_bot)
        count += 1
        if count == 25 : 
            num += 22
            count = 0 
        else:
            num += 1
   
    wb.close()

    return order.num

def OrderWorkExcelDownloadAPI(request):
    pk = request.POST.get('pk')

    file_name = WorkExcelCreate(pk)
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'file_name' : file_name
    }), content_type='application/json')

def OrdersWorkExcelDownloadAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    file_list = []
    path = './media/sales/work/'
    os.makedirs(path, exist_ok=True)
    for data in data_list: 
        file_name = WorkExcelCreate(data['id'])
        file_list.append(path + '/' + file_name + '.xlsx')
    if len(file_list) == 1 :
        return HttpResponse(
            json.dumps({
                'result': 'success',
                'sort' : 'xlsx',
                'file_name' : file_name
        }), content_type='application/json')
    else:
        date_te=arrow.now().format("YY-MM-DD-HHmm")
        zip_subdir = date_te
        zip_filename = "%s.zip" % zip_subdir
        os.makedirs(path, exist_ok=True)
        zf = zipfile.ZipFile(path+'/'+date_te+".zip", "w")

        for fpath in file_list:
            fdir, fname = os.path.split(fpath)
            zip_path = os.path.join(zip_subdir, fname)
            zf.write(fpath,zip_path)
        zf.close()
        file_name = str(zip_filename.split('.zip')[0])
        return HttpResponse(
            json.dumps({
                'result': 'success',
                'sort' : 'zip',
                'file_name' : file_name
        }), content_type='application/json')