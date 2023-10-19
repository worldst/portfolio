import json
from django.http import HttpResponse, JsonResponse
from system.models import Account
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q, F, Func, Value, CharField
from base.models import Code, Company, IntroducingCompany
from system.models import Account
from django.forms.models import model_to_dict
from sales.models import Estimate, EstimateFile
from datetime import datetime
from common.get_literal_eval import get_literal_eval
from ast import literal_eval
from common.revisions import set_revisions, get_revisions
from django.db import transaction
import openpyxl
import os
from openpyxl.styles import (
    Border, Side,
    Alignment,
	Font
)
@csrf_exempt
@transaction.atomic
def EstimateCreateAPI(request):
    json_data = json.loads(request.POST.get('obj'))
    
    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')

    try:
        company = Company.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )
    except Company.DoesNotExist:
        company = Company.objects.create(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )

    json_data['company_id'] = company.pk
    # 주문 특이사항 변환
    special_note = ''
    for row in json_data['special_note']:
        if special_note == '':
            special_note = row
        else: 
            special_note += 'Ψ' + row
    json_data['special_note'] = special_note

    num = datetime.now().strftime('E%Y%m%d-')
    estimate = Estimate.objects.filter(
        introducing_company_id = json_data['introducing_company_id'],
        num__icontains = num 
    ).order_by('-num')[0:1]

    if len(estimate) != 0 :
        estimate = estimate.first()
        estimate_len = len(estimate.num.split('-'))
        num = num + str(int(estimate.num.split('-')[estimate_len-1]) + 1 ).zfill(3)
    else:
        num = num + '1'.zfill(3)

    json_data['num'] = num
    estimate = Estimate.objects.create(**json_data)

    for upload_file in request.FILES.getlist('estimate_file'):
        EstimateFile.objects.create(
            estimate = estimate,
            upload_file = upload_file
        )
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def EstimateDelAPI(request,pk):

    Estimate.objects.filter(pk=pk).delete()

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def GetEstimatesAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    status = request.GET.get('status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Estimate.objects.filter(introducing_company=introducing_company).order_by('-created_dt')
    if status != '':
        data_list = data_list.filter(
            status = status
        )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(registration_date__icontains = keyword)|
            Q(pjt_name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'registration_date':
            q.add(Q(registration_date__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(pjt_name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','num','pjt_name','company_name','account__username','memo','status'
    ).annotate(
        created_dt  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        )
    )[(page-1) * list_range :  page * list_range]
    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)

@csrf_exempt
def GetEstimateAPI(request,pk):
    try:
        estimate = Estimate.objects.get(
            pk = pk
        )
        estimate_file = EstimateFile.objects.filter(
            estimate = estimate
        ).values('pk','upload_file')

        username = estimate.username
        estimate = (model_to_dict(estimate))
        estimate['username'] = username
        return JsonResponse(
            {
                'estimate': estimate,
                'estimate_file': list(estimate_file),
            }, safe=False)
    except Estimate.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')

@csrf_exempt
def EstimateModAPI(request, pk):
    
    json_data = json.loads(request.POST.get('obj'))
    
    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')

    try:
        company = Company.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )
    except Company.DoesNotExist:
        company = Company.objects.create(
            introducing_company_id = json_data['introducing_company_id'],
            name = json_data['company_name'],
        )

    json_data['company_id'] = company.pk
    # 주문 특이사항 변환
    special_note = ''
    for row in json_data['special_note']:
        if special_note == '':
            special_note = row
        else: 
            special_note += 'Ψ' + row
    json_data['special_note'] = special_note
    set_revisions(request, 'sales', 'estimate', pk, json_data)
    # Estimate.objects.filter(pk=pk).update(**json_data)

    estimate = Estimate.objects.get(
        pk = pk
    )
    EstimateFile.objects.filter(
        estimate = estimate
    ).delete()
    for upload_file in request.FILES.getlist('estimate_file'):
        EstimateFile.objects.create(
            estimate = estimate,
            upload_file = upload_file
        )

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')


@transaction.atomic
def EstimateListDelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')

    for data in data_list:
        estimate = Estimate.objects.get(
            id = data['id']
        )
        estimate.delete()
        
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')


@csrf_exempt
def EstimateStatusChangeAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    
    for data in data_list:
        estimate = Estimate.objects.get(
            id = data['id']
        )

        estimate.status = json_data['status']
        estimate.save()

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def EstimateExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    status = request.GET.get('status') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Estimate.objects.filter(introducing_company=introducing_company).order_by('-created_dt')
    if status != '':
        data_list = data_list.filter(
            status = status
        )
    if start_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__gte = start_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__gte = start_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'due_date':
            data_list = data_list.filter(
                due_date__lte = end_date
            )
        elif date_type == 'registration_date':
            data_list = data_list.filter(
                registration_date__lte = end_date
            )
        else:
            data_list = data_list.filter(
                created_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(num__icontains = keyword)|
            Q(registration_date__icontains = keyword)|
            Q(pjt_name__icontains = keyword)|
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'num':
            q.add(Q(num__icontains = keyword), q.AND)
        elif search_type == 'registration_date':
            q.add(Q(registration_date__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(pjt_name__icontains = keyword), q.AND)
        elif search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','num','pjt_name','company_name','account__username','memo','status'
    ).annotate(
        created_dt  = Func(
            F('created_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        )
    )

    path = './media/sales/estimate'
    docx_title = 'export_estimate'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_estimate.xlsx')
    ws = wb["Sheet 1"]
    i = 2

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)

    for data in data_list : 
        num = str(i)
        status = '견적대기'
        if data['status'] == 'CANCEL':
            status = '견적취소'
        if data['status'] == 'COMPLT':
            status = '견적확정'
        ws['A'+num].value = data['num']
        ws['B'+num].value = data['created_dt']
        ws['C'+num].value = data['pjt_name'] or ''
        ws['D'+num].value = data['company_name'] or ''
        ws['E'+num].value = data['account__username'] or ''
        ws['F'+num].value = data['memo'] or ''
        ws['G'+num].value = status

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font
        ws['F'+num].font = font
        ws['G'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
        ws['F'+num].border = thin_border
        ws['G'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
        ws['F'+num].alignment = alignment
        ws['G'+num].alignment = alignment

        ws.row_dimensions[i].height = 40

        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')
