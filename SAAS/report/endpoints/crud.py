import json
from django.http import HttpResponse, JsonResponse
from system.models import Account
from django.views.decorators.csrf import csrf_exempt
from django.forms.models import model_to_dict
from django.db.models import Sum, Q, F, Func, Value, CharField
from base.models import Company, Code, IntroducingCompany, Excel
from system.models import Account
from django.db import transaction
from common.revisions import set_revisions, get_revisions
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
import os
from openpyxl.styles import (
    Border, Side,
    Alignment,
	Font
)
from sales.models import ProcessRecord, Order
from dateutil.relativedelta import relativedelta
@csrf_exempt
def GetJobsReportAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)

    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = ProcessRecord.objects.filter(order_process__order__introducing_company=introducing_company,end_dt__isnull=False).order_by('-pk')

    if start_date != '':
        if date_type == 'create':
            data_list = data_list.filter(
                order_process__order__created_dt__date__gte = start_date
            )
        if date_type == 'start':
            data_list = data_list.filter(
                start_dt__date__gte = start_date
            )
        if date_type == 'end':
            data_list = data_list.filter(
                end_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'create':
            data_list = data_list.filter(
                order_process__order__created_dt__date__lte = end_date
            )
        if date_type == 'start':
            data_list = data_list.filter(
                start_dt__date__lte = end_date
            )
        if date_type == 'end':
            data_list = data_list.filter(
                end_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(order_process__order__company_name__icontains = keyword)|
            Q(order_process__order__pjt_name__icontains = keyword)|
            Q(order_process__process__icontains = keyword)|
            Q(worker__name__icontains = keyword)|
            Q(order_process__order__num__icontains = keyword)|
            Q(equipment_name__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'company_name':
            q.add(Q(order_process__order__company_name__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(order_process__order__pjt_name__icontains = keyword), q.AND)
        elif search_type == 'process':
            q.add(Q(order_process__process__icontains = keyword), q.AND)
        elif search_type == 'worker':
            q.add(Q(worker__name__icontains = keyword), q.AND)
        elif search_type == 'order_num':
            q.add(Q(order_process__order__num__icontains = keyword), q.AND)
        elif search_type == 'equipment':
            q.add(Q(equipment_name__icontains = keyword), q.AND)
            
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','order_process__order__num','order_process__order__company_name','order_process__order__pjt_name','equipment_name','worker__name'
    ).annotate(
        job_dt =   Func(
            F('end_dt') - F('start_dt'),
            Value('HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
        start_dt  = Func(
            F('start_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
        end_dt  = Func(
            F('end_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
    )[(page-1) * list_range :  page * list_range]

    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)

@csrf_exempt
def JobsExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    start_date = request.GET.get('start_date') or ''
    end_date = request.GET.get('end_date') or ''

    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = ProcessRecord.objects.filter(order_process__order__introducing_company=introducing_company,end_dt__isnull=False).order_by('-pk')

    if start_date != '':
        if date_type == 'create':
            data_list = data_list.filter(
                order_process__order__created_dt__date__gte = start_date
            )
        if date_type == 'start':
            data_list = data_list.filter(
                start_dt__date__gte = start_date
            )
        if date_type == 'end':
            data_list = data_list.filter(
                end_dt__date__gte = start_date
            )

    if end_date != '':
        if date_type == 'create':
            data_list = data_list.filter(
                order_process__order__created_dt__date__lte = end_date
            )
        if date_type == 'start':
            data_list = data_list.filter(
                start_dt__date__lte = end_date
            )
        if date_type == 'end':
            data_list = data_list.filter(
                end_dt__date__lte = end_date
            )

    if search_type == '':
        data_list = data_list.filter(
            Q(order_process__order__company_name__icontains = keyword)|
            Q(order_process__order__pjt_name__icontains = keyword)|
            Q(order_process__process__icontains = keyword)|
            Q(worker__name__icontains = keyword)|
            Q(order_process__order__num__icontains = keyword)|
            Q(equipment_name__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'company_name':
            q.add(Q(order_process__order__company_name__icontains = keyword), q.AND)
        elif search_type == 'pjt_name':
            q.add(Q(order_process__order__pjt_name__icontains = keyword), q.AND)
        elif search_type == 'process':
            q.add(Q(order_process__process__icontains = keyword), q.AND)
        elif search_type == 'worker':
            q.add(Q(worker__name__icontains = keyword), q.AND)
        elif search_type == 'order_num':
            q.add(Q(order_process__order__num__icontains = keyword), q.AND)
        elif search_type == 'equipment':
            q.add(Q(equipment_name__icontains = keyword), q.AND)
            
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','order_process__order__num','order_process__order__company_name','order_process__order__pjt_name','equipment_name','worker__name'
    ).annotate(
        job_dt = Func(
            F('end_dt') - F('start_dt'),
            Value('HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
        start_dt = Func(
            F('start_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
        end_dt  = Func(
            F('end_dt'),
            Value('YYYY-MM-DD HH24:MI:SS'),
            function='to_char',
            output_field=CharField()
        ),
    )
    path = './media/report'
    docx_title = 'export_job_report'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_job_report.xlsx')
    ws = wb["Sheet 1"]
    i = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)
    for data in data_list : 
        num = str(i)
        ws['A'+num].value = data['order_process__order__num']
        ws['B'+num].value = data['order_process__order__company_name'] or ''
        ws['C'+num].value = data['order_process__order__pjt_name'] or ''
        ws['D'+num].value = data['equipment_name'] or ''
        ws['E'+num].value = data['worker__name'] or ''
        ws['F'+num].value = data['start_dt'] or ''
        ws['G'+num].value = data['end_dt'] or ''
        ws['H'+num].value = data['job_dt'] or ''
     

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font
        ws['F'+num].font = font
        ws['G'+num].font = font
        ws['H'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
        ws['F'+num].border = thin_border
        ws['G'+num].border = thin_border
        ws['H'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
        ws['F'+num].alignment = alignment
        ws['G'+num].alignment = alignment
        ws['H'+num].alignment = alignment
       
        ws.row_dimensions[i].height = 40
     
        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')

@csrf_exempt
def GetSalesReportAPI(request):
    search_type = request.GET.get('search_type') or ''
    # ! 날짜형식은 필수
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    search_month = request.GET.get('search_month') or ''

    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Order.objects.filter(introducing_company=introducing_company).order_by('-pk')
    if search_type == '':
        data_list = data_list.filter(
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        data_list = data_list.filter(q)
    sales_list = []
    year = int(str(search_month).split('-')[0])
    month = int(str(search_month).split('-')[1])
    for i in range(1, 13):
        if date_type == 'create':
            data = Order.objects.filter(
                created_dt__year = year,
                created_dt__month = i,
            )
        if date_type == 'shipment':
            data = Order.objects.filter(
                shipment_end_dt__year = year,
                shipment_end_dt__month = i,
            )
        sales_list.append(
            dict(
                date = str(year) +'-' + str(i).zfill(2),
                sales_price = data.aggregate(
                    sales_price = Sum(
                        F('process_price') + F('material_price') + F('outsourcing_price') + F('etc_price')
                    )
                )['sales_price'] or 0 
            )
        )
    if search_month != '':
        date = datetime(year=year, month=month, day=1).date()
        ex_month = date - relativedelta(months=1)
        after_month = date + relativedelta(months = 1)
        if date_type == 'create':
            ex_year_list = data_list.filter(
                created_dt__year = year - 1,
                created_dt__month = month,
            )
            ex_month_list = data_list.filter(
                created_dt__date__gte = ex_month,
                created_dt__date__lt = date,
            )
            data_list = data_list.filter(
                created_dt__date__gte = date,
                created_dt__date__lt = after_month,
            )
        if date_type == 'shipment':
            ex_year_list = data_list.filter(
                shipment_end_dt__year = year - 1,
                shipment_end_dt__month = month,
            )
            ex_month_list = data_list.filter(
                shipment_end_dt__date__gte = ex_month,
                shipment_end_dt__date__lt = date,
            )
            data_list = data_list.filter(
                shipment_end_dt__date__gte = date,
                shipment_end_dt__date__lt = after_month,
            )
   
    data_len = len(data_list)
    all_sales_price, all_profit_loss, all_total_cost = 0, 0, 0
    all_process_cost, all_material_cost, all_outsourcing_cost = 0, 0, 0
    all_process_price, all_material_price, all_outsourcing_price, all_etc_price = 0, 0, 0, 0
    object_list = []
    
    date_list = []
    # company_key = []
    # company_data = []
    for data in data_list:
        # * 매출합계, 가공, 외주, 자재, 기타비용
        sales_price = data.total_sales_price
        process_price = data.process_price
        material_price = data.material_price
        outsourcing_price = data.outsourcing_price
        etc_price = data.etc_price
        # * 가공, 외주, 자재비 (원가)
        process_cost = data.process_cost
        material_cost = data.material_cost
        outsourcing_cost = data.outsourcing_cost
        # * 원가 총액
        total_cost = data.total_cost
        # * 손익
        profit_loss = sales_price - total_cost
        account = ''
        if data.account != None :
            account = data.account.name


        # 매출 계산
        all_sales_price += sales_price        
        all_process_price += process_price
        all_material_price += material_price
        all_outsourcing_price += outsourcing_price
        all_etc_price += etc_price

        # 원가 계산
        all_total_cost += total_cost
        all_process_cost += process_cost
        all_material_cost += material_cost
        all_outsourcing_cost += outsourcing_cost

        all_profit_loss += profit_loss
        
        object_list.append(
            dict(
                pk = data.pk,
                num = data.num,
                created_dt = data.created_dt.strftime('%Y-%m-%d %H:%M:%S'),
                pjt_name = data.pjt_name,
                company_name = data.company_name,
                account = account,
                sales_price = sales_price,
                profit_loss = profit_loss
            )
        )
        if date_type == 'create':
            date = data.created_dt.strftime('%Y-%m-%d')
        else:
            date = data.shipment_end_dt.strftime('%Y-%m-%d')

        date_list.append(
            dict(
                date = date,
                sales_price = sales_price
            )
        )
        # company_key.append(data.company_name)
        # company_data.append(
        #     dict(
        #         company_name = data.company_name,
        #         sales_price = sales_price
        #     )
        # )
    # * 업체별 현황
    # company_key = set(company_key)
   
    # company_list = []
    # for index, key in enumerate(company_key):
    #     company_list.append(dict(company_name=key, cnt=0, sales_price = 0))
    #     for data in company_data:
    #         if key == data['company_name']:
    #             company_list[index]['cnt'] += 1
    #             company_list[index]['sales_price'] += data['sales_price']
    
    ex_month_all_sales_price = ex_month_list.aggregate(
        all_sales_price = Sum('process_price') + Sum('material_price') + Sum('outsourcing_price') + Sum('etc_price')
    )['all_sales_price'] or 0
    # * 전월대비 증감
    ex_month_rate = 100 
    if ex_month_all_sales_price != 0 :
        ex_month_rate = int(((all_sales_price - ex_month_all_sales_price) / ex_month_all_sales_price) * 100)
    ex_year_all_sales_price = ex_year_list.aggregate(
        all_sales_price = Sum('process_price') + Sum('material_price') + Sum('outsourcing_price') + Sum('etc_price')
    )['all_sales_price'] or 0
    # * 전월대비 증감
    ex_month_rate = 100 
    if ex_month_all_sales_price != 0 :
        ex_month_rate = int(((all_sales_price - ex_month_all_sales_price) / ex_month_all_sales_price) * 100)
    # * 전년대비 증감
    ex_year_rate = 100 
    if ex_year_all_sales_price != 0 :
        ex_year_rate = int(((all_sales_price - ex_year_all_sales_price) / ex_year_all_sales_price) * 100)
    return JsonResponse({
        'data_len': data_len,
        'all_sales_price' : all_sales_price,
        'all_profit_loss' : all_profit_loss,
        'all_total_cost' : all_total_cost,
        'all_process_cost' : all_process_cost,
        'all_material_cost' : all_material_cost,
        'all_outsourcing_cost' : all_outsourcing_cost,
        'all_process_price' : all_process_price,
        'all_material_price' : all_material_price,
        'all_outsourcing_price' : all_outsourcing_price,
        'all_etc_price' : all_etc_price,
        'object_list': object_list,
        'sales_list' : sales_list,
        # 'company_list' : company_list,
        'ex_month_rate' : ex_month_rate,
        'ex_year_rate' : ex_year_rate
    }, safe=False)

@csrf_exempt
def GetSaleCompanysReportAPI(request):
    search_type = request.GET.get('search_type') or ''
    # ! 날짜형식은 필수
    date_type = request.GET.get('date_type') or ''
    keyword = request.GET.get('keyword') or ''
    start_month = request.GET.get('start_month') or ''
    end_month = request.GET.get('end_month') or ''

    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Order.objects.filter(introducing_company=introducing_company).order_by('-pk')
    if search_type == '':
        data_list = data_list.filter(
            Q(company_name__icontains = keyword)|
            Q(account__name__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'company_name':
            q.add(Q(company_name__icontains = keyword), q.AND)
        elif search_type == 'account':
            q.add(Q(account__name__icontains = keyword), q.AND)
        data_list = data_list.filter(q)

    if start_month != '':
        year = int(str(start_month).split('-')[0])
        month = int(str(start_month).split('-')[1])
        date = datetime(year=year, month=month, day=1).date()
        if date_type == 'create':
            data_list = data_list.filter(
                created_dt__date__gte = date,
            )
        if date_type == 'shipment':
            data_list = data_list.filter(
                shipment_end_dt__date__gte = date,
            )
    if end_month != '':
        year = int(str(end_month).split('-')[0])
        month = int(str(end_month).split('-')[1])
        date = datetime(year=year, month=month, day=1).date()
        after_month = date + relativedelta(months=1)
        if date_type == 'create':
            data_list = data_list.filter(
                created_dt__date__lt = after_month,
            )
        if date_type == 'shipment':
            data_list = data_list.filter(
                shipment_end_dt__date__lt = after_month,
            )
              
    company_key = []
    company_data = []
    total_sales_price = 0
    total_profit_loss = 0
    for data in data_list:
        #매출액
        sales_price = data.total_sales_price
        total_sales_price += sales_price
        #원가
        total_cost = data.total_cost
        # 손익
        profit_loss = sales_price - total_cost
        company_key.append(data.company_name)
        total_profit_loss += profit_loss
        company_data.append(
            dict(
                company_name = data.company_name,
                sales_price = sales_price,
                total_cost = total_cost,
                profit_loss = profit_loss
            )
        )
    # * 업체별 현황
    company_key = set(company_key)
   
    company_list = []
    for index, key in enumerate(company_key):
        company_list.append(dict(company_name=key, cnt=0, sales_price = 0, total_cost = 0, profit_loss = 0, rate = 0))
        for data in company_data:
            if key == data['company_name']:
                company_list[index]['cnt'] += 1
                company_list[index]['sales_price'] += data['sales_price']
                company_list[index]['total_cost'] += data['total_cost']
                company_list[index]['profit_loss'] += data['profit_loss']
    
    for company in company_list:
        if total_sales_price != 0 and company['sales_price'] != 0 :
            company['rate'] = round(company['sales_price']  / total_sales_price * 100, 2) 

    company_list_by_rate = sorted(company_list, key=lambda x:(x['sales_price']),reverse=True)
    company_list_by_profit = sorted(company_list, key=lambda x: (x['profit_loss']), reverse=True)
    
    return JsonResponse({
        'data_len': len(company_list),
        'company_list_by_rate' : company_list_by_rate,
        'company_list_by_profit' : company_list_by_profit,
    }, safe=False)
