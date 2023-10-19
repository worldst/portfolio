import json
from django.db.models import Q, F
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.hashers import check_password, make_password
from django.views.decorators.csrf import csrf_exempt
from common.get_literal_eval import get_literal_eval
from common.mail import send
from system.models import Account
from base.models import Code, Equipment, IntroducingCompany
from django.contrib.auth.hashers import check_password, make_password
from django.db import transaction
from common.revisions import set_revisions, get_revisions
import xlsxwriter
import qrcode
import os
import zipfile
import arrow
import openpyxl
from openpyxl.styles import (
    Border, Side,
    Alignment,
	Font
)
def get_random_pwd():
    import string
    import random

    new_pw_len = 10 # 새 비밀번호 길이

    print("영어대소문자", string.ascii_letters)
    print("숫자", string.digits)
    #print("특수문자", string.punctuation)

    pw_candidate = string.ascii_letters + string.digits + string.punctuation 

    new_pw = ""
    for i in range(new_pw_len):
        new_pw += random.choice(pw_candidate)

    print("생성된 랜덤 비밀번호", new_pw)

    return new_pw


@csrf_exempt
def ChkPasswordEndpoint(request, pk):
    """
    사용자 비밀번호 확인 API
    """
    print(request.POST.get('password'))
    print(pk)

    user = Account.objects.get(pk=pk)

    print(user.password)

    if check_password(request.POST.get('password'), user.password):
        status = 'success'
    else:
        status = 'fail'

    return HttpResponse(
        json.dumps({
        'result': status
    }), content_type='application/json')


@csrf_exempt
def ModAccountEndpoint(request, pk):
    """
    사용자 수정 API
    """
    user = Account.objects.get(pk=pk)

    user.password = make_password(request.POST.get('password'))
    user.save()
    
    return HttpResponse(
        json.dumps({
        'result': 'success'
    }), content_type='application/json')


@csrf_exempt
def FindAccountEndpoint(request):
    """
    사용자 아이디/비밀번호 찾기 API
    """
    result = 'None'
    if request.POST.get('type') == 'id':
        if Account.objects.filter(email=request.POST.get('email')).exists():
            account_obj = Account.objects.get(email=request.POST.get('email'))

            send(
                '아이디 전송드립니다.',
                '아이디는 ' + account_obj.username  + ' 입니다.',
                account_obj.email
            )
        else:
            print('존재하지 않는 이메일입니다.')
            result = 'A'
    else:
        if Account.objects.filter(username=request.POST.get('id'),email=request.POST.get('email')).exists():
            new_pwd = get_random_pwd()
            account_obj = Account.objects.get(username=request.POST.get('id'),email=request.POST.get('email'))
            account_obj.password = make_password(new_pwd)
            account_obj.save()

            send(
                '비밀번호 전송드립니다.',
                '임시 비밀번호는 ' + new_pwd  + ' 입니다.',
                account_obj.email
            )
        else:
            print('존재하지 않는 이메일 혹은 아이디입니다.')
            result = 'B'

    return HttpResponse(
        json.dumps({
            'result': result
    }), content_type='application/json')


def GetLoginAccountEndpoint(request):
    """
    현재 로그인 된 사용자 목록 확인
    """
    return JsonResponse(list(Account.objects.filter(is_login_checked=True).values('id','username','type','factory','process')), safe=False)


@csrf_exempt
def ProcessEquipmentAPI(request):

    """
    공정 장비 변경
    """
     
    json_data = json.loads(request.body.decode('utf-8'))
    # print(request.session['username'])
    try:
        account = Account.objects.get(
            introducing_company_id = request.session['introducing_company_id'],
            username = request.session['username']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_account'
        }), content_type='application/json')
    
    '''
    try:
        process = Code.objects.get(
            pk = int(json_data.pop('process'))
        )
    except Code.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'no_process'
        }), content_type='application/json')
    try:
        equipment = Equipment.objects.get(
            pk = int(json_data.pop('equipment'))
        )
    except Equipment.DoesNotExist:
        return HttpResponse(
            json.dumps({
                'result': 'no_process'
        }), content_type='application/json')
    '''

    account.process_id = json_data.pop('process')
    account.equipment_id = json_data.pop('equipment')
    account.save()

    return HttpResponse(
        json.dumps({
            'result': 'success',
    }), content_type='application/json')

@csrf_exempt
def AccountCreateAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    try:
        json_data['introducing_company_id'] = request.session['introducing_company_id']
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    try:
        Account.objects.get(
            introducing_company_id = json_data['introducing_company_id'],
            username = json_data['username'],
        )
        return HttpResponse(
            json.dumps({
                'result': 'username'
        }), content_type='application/json')
    except Account.DoesNotExist:
        try:
            Account.objects.get(
                introducing_company_id = json_data['introducing_company_id'],
                name = json_data['name'],
                is_delete = False
            )
            return HttpResponse(
                json.dumps({
                    'result': 'name'
            }), content_type='application/json')
        except Account.DoesNotExist:
            json_data['password'] = make_password(json_data['password'])
            Account.objects.create(**json_data)
        except Account.MultipleObjectsReturned:
            return HttpResponse(
                json.dumps({
                    'result': 'name'
            }), content_type='application/json')

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def GetAccountsAPI(request):
    search_type = request.GET.get('search_type') or ''
    keyword = request.GET.get('keyword') or ''
    #페이지당 데이터 기본 10개
    list_range = request.GET.get('list_range') or 10
    list_range = int(list_range)
    #페이지 기본 1페이지
    page = request.GET.get('page') or 1
    page = int(page)


    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Account.objects.filter(~Q(type="DK"), introducing_company=introducing_company, is_delete=False).order_by('-pk')

    if search_type == '':
        data_list = data_list.filter(
            Q(username__icontains = keyword)|
            Q(name__icontains = keyword)|
            Q(type__icontains = keyword)|
            Q(email__icontains = keyword)|
            Q(phone_number__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'username':
            q.add(Q(username__icontains = keyword), q.AND)
        elif search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'type':
            q.add(Q(type__icontains = keyword), q.AND)
        elif search_type == 'email':
            q.add(Q(email__icontains = keyword), q.AND)
        elif search_type == 'phone_number':
            q.add(Q(phone_number__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
            
        data_list = data_list.filter(q)

    data_len = len(data_list)
    data_list = data_list.values(
        'pk','username','name','type','email','phone_number','memo','is_active'
    )[(page-1) * list_range :  page * list_range]

    return JsonResponse({
        'data_len': data_len,
        'data_list':list(data_list),
    }, safe=False)


@transaction.atomic
def AccountsDelAPI(request):
    json_data = json.loads(request.body.decode('utf-8'))
    data_list = json_data.pop('data_list')
    for data in data_list:
        account = Account.objects.get(
            id = data['id']
        )
        account.is_delete = True
        account.save()
        # account.delete()
    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def ModAccountAPI(request,pk):
    json_data = json.loads(request.body.decode('utf-8'))

    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    try:
        Account.objects.get(
            ~Q(pk = pk),
            introducing_company = introducing_company,
            name = json_data['name'],
            is_delete = False
        )
        return HttpResponse(
            json.dumps({
                'result': 'fail'
        }), content_type='application/json')
    except Account.DoesNotExist:
        json_data['password'] = make_password(json_data['password'])
        set_revisions(request, 'system', 'account', pk, json_data)
    except Account.MultipleObjectsReturned:
        return HttpResponse(
            json.dumps({
                'result': 'name'
        }), content_type='application/json')

    return HttpResponse(
        json.dumps({
            'result': 'success'
    }), content_type='application/json')

@csrf_exempt
def AccountExcelDownloadAPI(request):
    search_type = request.GET.get('search_type') or ''
    keyword = request.GET.get('keyword') or ''
    try:
        introducing_company = IntroducingCompany.objects.get(
            pk = request.session['introducing_company_id']
        )
    except:
        return HttpResponse(
            json.dumps({
                'result': 'no_session'
        }), content_type='application/json')
    
    data_list = Account.objects.filter(~Q(type="DK"), introducing_company=introducing_company, is_delete=False).order_by('-pk')

    if search_type == '':
        data_list = data_list.filter(
            Q(username__icontains = keyword)|
            Q(name__icontains = keyword)|
            Q(email__icontains = keyword)|
            Q(phone_number__icontains = keyword)|
            Q(memo__icontains = keyword)
        )
    else:
        q = Q()
        if search_type == 'username':
            q.add(Q(username__icontains = keyword), q.AND)
        elif search_type == 'name':
            q.add(Q(name__icontains = keyword), q.AND)
        elif search_type == 'email':
            q.add(Q(email__icontains = keyword), q.AND)
        elif search_type == 'phone_number':
            q.add(Q(phone_number__icontains = keyword), q.AND)
        elif search_type == 'memo':
            q.add(Q(memo__icontains = keyword), q.AND)
            
        data_list = data_list.filter(q)

    data_list = data_list.values(
        'pk','username','name','email','phone_number','memo','is_active'
    )

    path = './media/system/'
    docx_title = 'export_account'
    os.makedirs(path, exist_ok=True)
        
    wb = openpyxl.load_workbook('./main/static/excel_file/export_account.xlsx')
    ws = wb["Sheet 1"]
    i = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='"SUIT Variable", sans-serif', size=11, bold=True)
    for data in data_list : 
        num = str(i)
        is_active = 'X'
        if data['is_active'] :
            is_active = 'O'
        ws['A'+num].value = data['username']
        ws['B'+num].value = data['name'] or ''
        ws['C'+num].value = data['email'] or ''
        ws['D'+num].value = data['phone_number'] or ''
        ws['E'+num].value = data['memo'] or ''
        ws['F'+num].value = is_active
     

        ws['A'+num].font = font
        ws['B'+num].font = font
        ws['C'+num].font = font
        ws['D'+num].font = font
        ws['E'+num].font = font
        ws['F'+num].font = font

        ws['A'+num].border = thin_border
        ws['B'+num].border = thin_border
        ws['C'+num].border = thin_border
        ws['D'+num].border = thin_border
        ws['E'+num].border = thin_border
        ws['F'+num].border = thin_border
   
        ws['A'+num].alignment = alignment
        ws['B'+num].alignment = alignment
        ws['C'+num].alignment = alignment
        ws['D'+num].alignment = alignment
        ws['E'+num].alignment = alignment
        ws['F'+num].alignment = alignment
       
        ws.row_dimensions[i].height = 40
     
        i +=1
    wb.save(path + '/' + docx_title + '.xlsx')
    wb.close()
    return HttpResponse(
        json.dumps({
            'result': 'success',
            'name': docx_title
        }), content_type='application/json')
